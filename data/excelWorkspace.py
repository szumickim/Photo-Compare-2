import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os
from consts import *
import datetime
from tkinter import messagebox

LAST_COLUMN_IN_EXPORT: int = 40
DIFFERENT: int = 2
SAME: int = 1

IS_DIFFERENT = "Different"
IS_SIMILAR = "Similar"
POSSIBLE: str = "Possible"
MANUAL: str = "Similar - manual"

ALL_IMAGES: str = "Show all images"
COMPARE: str = "Compare"

SIMILARITY_TYPE_COLUMN: str = "Similarity type"

PRODUCTS_ON_SCREEN: int = 3

SHOW_ALL_SUMMARY_FORMAT: int = 1
SUMMARY_FORMAT: int = 2


def format_excel(final_df, file_name, format_type):
    # Stworzenie pustego obiektu xlsx
    writer = pd.ExcelWriter(file_name)

    # Formatowanie (tutaj dzieje się magia)
    format_sheet(writer, "sheet1", final_df, format_type)

    # Zamknięcie obiektu ze sformatowanym raportem
    writer.close()


def format_sheet(writer, sheet_name, df, format_type):
    # Zapisanie df do Excela
    df.to_excel(writer, sheet_name=sheet_name, index=False, freeze_panes=(1, 0))

    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    (max_row, max_col) = df.shape

    # formatowanie headerów
    header_format = workbook.add_format({
        'valign': 'vcenter',
        'bold': True,
        'text_wrap': True,
        'align': 'center',
        'border': 1})

    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num, value, header_format)

    worksheet.autofilter(0, 0, max_row, max_col - 1)  # nakładanie filtra

    green_format = workbook.add_format({
        'fg_color': '#EBF1DE'})
    salmon_format = workbook.add_format({
        'fg_color': '#FDE9D9'})
    red_format = workbook.add_format({
        'fg_color': '#E6B8B7'})

    if format_type == SUMMARY_FORMAT:
        green_columns_list = [1, 5, 9]
        salmon_columns_list = [2, 6]
        red_columns_list = [0]
    else:
        green_columns_list = [1]
        salmon_columns_list = [2]
        red_columns_list = [0]

    # Auto-adjust columns' width
    for column in df:

        column_width = df[column].astype(str).map(len).max()

        if np.isnan(column_width) or column_width < 12:
            column_width = 12
        col_idx = df.columns.get_loc(column)
        if col_idx in green_columns_list:
            worksheet.set_column(col_idx, col_idx, column_width, green_format)
        elif col_idx in salmon_columns_list:
            worksheet.set_column(col_idx, col_idx, column_width, salmon_format)
        elif col_idx in red_columns_list:
            worksheet.set_column(col_idx, col_idx, column_width, red_format)
        else:
            worksheet.set_column(col_idx, col_idx, column_width)
    # Kolorowanie kolumn
    # green_format = workbook.add_format({
    #     'fg_color': '#DCE6F1'})
    # worksheet.set_column(7, 11, 50, green_format)


def create_summary_excel(products_collection, folder_to_save):
    summary_columns = [SummaryConst.PRODUCT_ID, SummaryConst.FIRST_PHOTO_ID, SummaryConst.FIRST_REFERENCE,
                       SummaryConst.HEIGHT,
                       SummaryConst.WIDTH, SummaryConst.SECOND_PHOTO_ID, SummaryConst.SECOND_REFERENCE,
                       SummaryConst.SECOND_HEIGHT, SummaryConst.SECOND_WIDTH, SummaryConst.WORSE]
    df_summary = pd.DataFrame(columns=summary_columns)

    for product in products_collection:
        if product.photos_connection_type in ([IS_SIMILAR, MANUAL, POSSIBLE]):
            for photos_pair in product.photos_pairs_list:
                first_photo = photos_pair.photo1
                second_photo = photos_pair.photo2
                better_photo = photos_pair.better_photo if photos_pair.better_photo else ""

                df_summary.loc[df_summary.shape[0]] = \
                    [getattr(product, "<ID>"), first_photo.name, first_photo.asset_type,
                     first_photo.height, first_photo.width, second_photo.name, second_photo.asset_type,
                     second_photo.height, second_photo.width, better_photo]

    # format_excel(df_summary, f"{folder_to_save}/Summary {set_today_date('-')}.xlsx", SUMMARY_FORMAT)
    df_summary.to_excel(f"{folder_to_save}/Summary {set_today_date('-')}.xlsx", index=False)


def add_column_to_excel(excel_path, products_collection, first_row):
    workbook = load_workbook(filename=excel_path)
    sheet = workbook.active

    last_column = 1
    while sheet.cell(row=1, column=last_column).value and sheet.cell(row=1,
                                                                     column=last_column).value != "Similarity type":
        last_column += 1

    if first_row == 0:
        sheet.cell(row=1, column=last_column).value = SIMILARITY_TYPE_COLUMN
        sheet.cell(row=1, column=last_column + 1).value = "Similar pairs"

    start_row = first_row + 2
    i = 0
    for product in products_collection:
        sheet.cell(row=start_row + i, column=last_column).value = product.photos_connection_type
        sheet.cell(row=start_row + i, column=last_column + 1).value = ";" \
            .join(f"{photos_pair.similarity_type},{photos_pair.photo1.name},{photos_pair.photo2.name}"
                  for photos_pair in product.photos_pairs_list)
        i += 1
    # save the file
    workbook.save(filename=excel_path)


def work_with_show_all_summ_excel(products_collection, entry_info):
    folder_to_save = entry_info.photo_path if len(entry_info.photo_path) > 0 else os.path.dirname(entry_info.excel_path)

    excel_path = f"{folder_to_save}/STEPShowAllSummary {set_today_date('-')}.xlsx" if entry_info.data_from_step \
        else f"{folder_to_save}/ShowAllSummary {set_today_date('-')}.xlsx"

    to_modify = is_summary_excel_exists(excel_path)

    if to_modify:
        modify_show_all_summ_excel(products_collection, excel_path, entry_info)
    else:
        create_show_all_summ_excel(products_collection, excel_path, entry_info)


def is_summary_excel_exists(excel_path):
    if os.path.isfile(excel_path):
        if messagebox.askyesno("Summary excel",
                                  "Summary excel exists! Click 'YES' to add new values to existing file "
                                  "or 'NO' to overwrite older excel."):
            return True
    return False


def backup_excel(products_collection, entry_info):
    folder_to_save = entry_info.photo_path if len(entry_info.photo_path) > 0 else os.path.dirname(entry_info.excel_path)

    excel_path = f"{folder_to_save}/Backup{set_today_date('-')}.xlsx" if entry_info.data_from_step \
        else f"{folder_to_save}/Backup{set_today_date('-')}.xlsx"

    create_show_all_summ_excel(products_collection, excel_path, entry_info)


def get_all_summary_columns(entry_info):
    if entry_info.data_from_step:
        return [ShowAllConst.ASSET_ID, ShowAllConst.PRODUCT_IDS, ShowAllConst.PHOTO_REFERENCE_TYPE]
    else:
        return [ShowAllConst.PRODUCT_ID, ShowAllConst.PHOTO_ID, ShowAllConst.PHOTO_REFERENCE,
                ShowAllConst.HEIGHT, ShowAllConst.WIDTH]


def create_show_all_summ_excel(products_collection, excel_path, entry_info):
    show_all_summ_columns = get_all_summary_columns(entry_info)

    df_show_all_summ = pd.DataFrame(columns=show_all_summ_columns)

    for product in products_collection:
        for photo in product.all_photos:
            if photo.selected_photo:
                df_show_all_summ.loc[df_show_all_summ.shape[0]] = product_data_to_summary_excel(product, photo, entry_info)

    df_show_all_summ.to_excel(excel_path, index=False)


def product_data_to_summary_excel(product, photo, entry_info):
    return [photo.name, product.product_id, photo.asset_type] if entry_info.data_from_step else \
        [product.product_id, photo.name, photo.asset_type, photo.height, photo.width]


def modify_show_all_summ_excel(products_collection, excel_path, entry_info):
    workbook = load_workbook(filename=excel_path)
    sheet = workbook.active

    i = 1
    while sheet.cell(row=i, column=1).value:
        i += 1

    counter = i
    for product in products_collection:
        for photo in product.all_photos:
            if photo.selected_photo:
                for column, element in enumerate(product_data_to_summary_excel(product, photo, entry_info), start=1):
                    sheet.cell(row=counter, column=column).value = element

                counter += 1

    workbook.save(filename=excel_path)


def set_today_date(date_separator):
    date_time = datetime.datetime.now()
    return date_time.strftime(f"%d{date_separator}%m{date_separator}%Y")
